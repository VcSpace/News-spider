import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
#from requests.cookies import RequestsCookieJar
import threading
import time
import re

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
    #'Cookie': 'xqat=3e14cc861fdd960a5d84e7316165286b1bfeafe3;',
}

class SelfStock(object):
    threadLock = threading.Lock()
    def __init__(self, file_name):
        self.xlsxname = file_name

    def Style(self):
        self.m_font = Font(
            size=12,
            bold=True,
        )
        self.head_font = Font(
            size=14,
            bold=True,
        )

    def Deal_Xq_quote(self, data, name):
        self.threadLock.acquire()
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet(name)
        t_row = 1
        t_col = 1

        data_json = data['data']
        data_items = data_json['items']
        data_mk = data_items[0]['market']
        data_quote = data_items[0]['quote']
        #print(data_quote)

        sheet.cell(row=t_row, column=t_col, value="股票代码")
        sheet.cell(row=t_row, column=t_col + 1, value="股票名称")
        sheet.cell(row=t_row, column=t_col + 2, value="交易状态")
        sheet.cell(row=t_row, column=t_col + 3, value="更新时间")
        t_row = t_row + 1
        m_status = data_mk['status']
        stock_code = data_quote['symbol']
        stock_name = data_quote['name']
        m_time = data_quote['timestamp']
        timeStamp = float(m_time / 1000) #13位时间戳
        timeArray = time.localtime(timeStamp)
        otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

        sheet.cell(row=t_row, column=t_col, value=stock_code)
        sheet.cell(row=t_row, column=t_col + 1, value=stock_name)
        sheet.cell(row=t_row, column=t_col + 2, value=m_status)
        sheet.cell(row=t_row, column=t_col + 3, value=otherStyleTime)
        t_row = t_row + 2

        sheet.cell(row=t_row, column=t_col + 0, value="当前价格")
        sheet.cell(row=t_row, column=t_col + 1, value="涨跌幅度")
        sheet.cell(row=t_row, column=t_col + 2, value="涨跌价格")
        sheet.cell(row=t_row, column=t_col + 3, value="开盘价格")
        sheet.cell(row=t_row, column=t_col + 4, value="当前新高")
        sheet.cell(row=t_row, column=t_col + 5, value="当前新低")
        sheet.cell(row=t_row, column=t_col + 6, value="昨日收盘")
        sheet.cell(row=t_row, column=t_col + 7, value="平均价格")
        sheet.cell(row=t_row, column=t_col + 8, value="涨停价格")
        sheet.cell(row=t_row, column=t_col + 9, value="跌停价格")
        sheet.cell(row=t_row, column=t_col + 10, value="52周最高")
        sheet.cell(row=t_row, column=t_col + 11, value="52周最低")
        t_row = t_row + 1

        m_current = data_quote['current']
        m_percent = data_quote['percent'] #涨跌幅度
        m_chg = data_quote['chg'] #涨跌价格
        m_open = data_quote['open'] #开盘价
        m_yesclose = data_quote['last_close'] #昨收
        m_high = data_quote['high']
        m_low = data_quote['low']
        m_avg_price = data_quote['avg_price']
        m_limit_up = data_quote['limit_up'] #涨停
        m_limit_down = data_quote['limit_down'] #跌停
        m_high52w = data_quote['high52w'] #52周最高
        m_low52w = data_quote['low52w'] #52周最低
        sheet.cell(row=t_row, column=t_col + 0, value=m_current)
        sheet.cell(row=t_row, column=t_col + 1, value=m_percent)
        sheet.cell(row=t_row, column=t_col + 2, value=m_chg)
        sheet.cell(row=t_row, column=t_col + 3, value=m_open)
        sheet.cell(row=t_row, column=t_col + 4, value=m_high)
        sheet.cell(row=t_row, column=t_col + 5, value=m_low)
        sheet.cell(row=t_row, column=t_col + 6, value=m_yesclose)
        sheet.cell(row=t_row, column=t_col + 7, value=m_avg_price)
        sheet.cell(row=t_row, column=t_col + 8, value=m_limit_up)
        sheet.cell(row=t_row, column=t_col + 9, value=m_limit_down)
        sheet.cell(row=t_row, column=t_col + 10, value=m_high52w)
        sheet.cell(row=t_row, column=t_col + 11, value=m_low52w)
        t_row = t_row + 2


        sheet.cell(row=t_row, column=t_col + 0, value="成交额(/万)")
        sheet.cell(row=t_row, column=t_col + 1, value="成交量(/万手)")
        sheet.cell(row=t_row, column=t_col + 2, value="换手率")
        sheet.cell(row=t_row, column=t_col + 3, value="量比")
        sheet.cell(row=t_row, column=t_col + 4, value="振幅")
        sheet.cell(row=t_row, column=t_col + 5, value="市盈率TTM")
        sheet.cell(row=t_row, column=t_col + 6, value="市盈率(动)")
        sheet.cell(row=t_row, column=t_col + 7, value="市盈率(静)")
        sheet.cell(row=t_row, column=t_col + 8, value="市净率")
        sheet.cell(row=t_row, column=t_col + 9, value="总股本(/万)")
        sheet.cell(row=t_row, column=t_col + 10, value="流通股(/万)")
        sheet.cell(row=t_row, column=t_col + 11, value="总市值(/亿)")
        sheet.cell(row=t_row, column=t_col + 12, value="流通市值(/亿)")
        t_row = t_row + 1

        m_amount = data_quote['amount'] #成交额
        m_turnover_rate = data_quote['turnover_rate'] #换手：0.74%
        m_volume = data_quote['volume'] #成交量
        m_amplitude = data_quote['amplitude'] #振幅
        m_total_shares = data_quote['total_shares'] #总股本
        m_float_shares = data_quote['float_shares'] #流通股
        m_volume_ratio = data_quote['volume_ratio'] #量比
        m_pe_ttm = data_quote['pe_ttm'] #市盈率
        m_pe_forecast = data_quote['pe_forecast'] #动态市盈率
        m_pe_lyr = data_quote['pe_lyr'] #静态市盈率
        m_pb = data_quote['pb'] #市净率
        m_profit = data_quote['profit'] #年报利润
        m_profit_four = data_quote['profit_four'] #也是利润 不清楚
        m_profit_forecast = data_quote['profit_forecast']
        m_market_capital = data_quote['market_capital'] #总市值
        m_float_market_capital = data_quote['float_market_capital'] #流通市值
        sheet.cell(row=t_row, column=t_col + 0, value=round(m_amount / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 1, value=round(m_volume / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 2, value=str(m_turnover_rate) + "%")
        sheet.cell(row=t_row, column=t_col + 3, value=m_volume_ratio)
        sheet.cell(row=t_row, column=t_col + 4, value=str(m_amplitude) + "%")
        sheet.cell(row=t_row, column=t_col + 5, value=m_pe_ttm)
        sheet.cell(row=t_row, column=t_col + 6, value=m_pe_forecast)
        sheet.cell(row=t_row, column=t_col + 7, value=m_pe_lyr)
        sheet.cell(row=t_row, column=t_col + 8, value=m_pb)
        sheet.cell(row=t_row, column=t_col + 9, value=round(m_total_shares / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 10, value=round(m_float_shares / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 11, value=round(m_market_capital / 100000000, 2))
        sheet.cell(row=t_row, column=t_col + 12, value=round(m_float_market_capital / 100000000, 2))
        try:
            wb.save(self.xlsxname)
            self.threadLock.release()
        except Exception:
            print("Self_Stock Save Error = Xq_quote")
            self.threadLock.release()
    
    def Deal_Xq_distribution(self, data, name):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name(name)
        t_row = sheet.max_row + 2
        t_col = 1

        m_text = data['data']['analysis'][0] #今日主力净流入XX亿
        sheet.cell(row=t_row, column=t_col, value=m_text)
        t_row = t_row + 1

        sheet.cell(row=t_row, column=t_col, value="主力资金(/万)")
        sheet.cell(row=t_row, column=6, value="净流入(/万)")
        t_row = t_row + 1

        m_data = data['data']['distribution']

        m_sell = m_data['sell']
        m_buy = m_data['buy']

        sheet.cell(row=t_row, column=t_col, value="特大单卖出")
        sheet.cell(row=t_row + 1, column=t_col, value="大单卖出")
        sheet.cell(row=t_row + 2, column=t_col, value="中单卖出")
        sheet.cell(row=t_row + 3, column=t_col, value="小单卖出")
        sheet.cell(row=t_row + 4, column=t_col, value="合计")
        se_xlarge = m_sell['xlarge']
        se_large = m_sell['large']
        se_medium = m_sell['medium']
        se_small = m_sell['small']
        sum_sell = se_large + se_large + se_medium + se_small

        by_xlarge = m_buy['xlarge']
        by_large = m_buy['large']
        by_medium = m_buy['medium']
        by_small = m_buy['small']
        sum_buy = by_xlarge + by_large + by_medium + by_small
        t_col = t_col + 1
        sheet.cell(row=t_row, column=t_col, value=round(se_xlarge / 10000, 2))
        sheet.cell(row=t_row + 1, column=t_col, value=round(se_large / 10000, 2))
        sheet.cell(row=t_row + 2, column=t_col, value=round(se_medium / 10000, 2))
        sheet.cell(row=t_row + 3, column=t_col, value=round(se_small / 10000, 2))
        sheet.cell(row=t_row + 4, column=t_col, value=round(sum_sell / 10000, 2))
        t_col = t_col + 2

        sheet.cell(row=t_row, column=t_col, value=round(by_xlarge / 10000, 2))
        sheet.cell(row=t_row + 1, column=t_col, value=round(by_large / 10000, 2))
        sheet.cell(row=t_row + 2, column=t_col, value=round(by_medium / 10000, 2))
        sheet.cell(row=t_row + 3, column=t_col, value=round(by_small / 10000, 2))
        sheet.cell(row=t_row + 4, column=t_col, value=round(sum_buy / 10000, 2))
        t_col = t_col + 1

        sheet.cell(row=t_row, column=t_col, value="特大单买入")
        sheet.cell(row=t_row + 1, column=t_col, value="大单买入")
        sheet.cell(row=t_row + 2, column=t_col, value="中单买入")
        sheet.cell(row=t_row + 3, column=t_col, value="小单买入 ")
        sheet.cell(row=t_row + 4, column=t_col, value="净额 ")

        m_xlarge = round((by_xlarge / 10000) - (se_xlarge / 10000), 2)
        m_large = round((by_large / 10000) - (se_large / 10000), 2)
        m_medium = round((by_medium / 10000) - (se_medium / 10000), 2)
        m_small = round((by_small / 10000) - (se_small / 10000), 2)
        sheet.cell(row=t_row, column=6, value=m_xlarge)
        sheet.cell(row=t_row + 1, column=6, value=m_large)
        sheet.cell(row=t_row + 2, column=6, value=m_medium)
        sheet.cell(row=t_row + 3, column=6, value=m_small)
        sheet.cell(row=t_row + 4, column=6, value=round(sum_buy - sum_sell, 2) / 10000)
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Self_Stock Save Error = distribution")

    def Deal_Xq_query(self, data, name):
        print(2)

    con = 0
    def Deal_Xq(self, data, name):
        con = self.con
        if con < 3:
            if con == 0:
                con = con + 1
                t1 = threading.Thread(target=self.Deal_Xq_quote, args=(data, name, ))
                t1.start()
                t1.join()
            elif con == 1:
                con = con + 1
                t2 = threading.Thread(target=self.Deal_Xq_distribution, args=(data, name, ))
                t2.start()
                t2.join()
            elif con == 2:
                con = con + 1
                t3 = threading.Thread(target=self.Deal_Xq_query, args=(data, name, ))
                t3.start()
                t3.join()
        #elif con < 5:
        self.con = self.con + 1


    def get_SelfStock(self):
        #url_list = list()
        name_list = dict()
        t = time.time()
        m_time = int(t)
        with open("Code.txt", "r") as f:
            for line in f.readlines():
                m_line = line.strip('\n')  # 去掉列表中每一个元素的换行符
                sep = '#'
                code = line.split(sep, 1)[0]
                if code != '':
                    name = m_line.split(sep)[1]
                    m_code = re.sub('[a-zA-Z]', "", code)
                    url_code = 'https://stock.xueqiu.com/v5/stock/batch/quote.json?extend=detail&is_delay_ft=1&is_delay_hk=0&symbol={}'.format(code)
                    url_mainc = 'https://stock.xueqiu.com/v5/stock/capital/distribution.json?symbol={}&_={}'.format(code, m_time) #今日流出
                    url_main_h = 'https://stock.xueqiu.com/v5/stock/capital/query.json?count=20&symbol={}&_={}'.format(code, m_time) #流出历史
                    name_list.setdefault(name, [])
                    name_list[name].append(url_code)
                    name_list[name].append(url_mainc)
                    name_list[name].append(url_main_h)

        url = 'https://xueqiu.com'
        session = requests.session()
        session.get(url, headers=headers)
        for name in name_list:
            self.con = 0
            for m_url in name_list[name]:
                res = "xueqiu" in m_url
                if res == True:
                    resp = session.get(m_url, headers=headers)
                    data = json.loads(resp.text)
                    self.Deal_Xq(data, name)

            """
            resp = session.get(m_url, headers=headers)
            data = json.loads(resp.text)
            t1 = threading.Thread(target=self.Deal_Xq_data, args=(data, url, name, ))
            #t2 = threading.Thread(target=self.get_Main_capital_history, args=(data, url, name, ))
            t1.start()
            t1.join()
            break
            """
        del name_list #释放
        self.con = 0

    def main(self, file_name):
        Stock = SelfStock(file_name)
        Stock.get_SelfStock()
